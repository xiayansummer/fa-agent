from __future__ import annotations
import json
import httpx
from harness.skill_registry import skill, skill_registry
from config import settings

BASE_URL = "https://qimingpianapi.investarget.com"


def _base(extra: dict | None = None) -> dict:
    data = {"open_id": settings.qmingpian_token}
    if extra:
        data.update(extra)
    return data


def _check(resp: httpx.Response) -> dict:
    data = resp.json()
    if str(data.get("status")) != "0":
        raise ValueError(f"企名片 API error {data.get('status')}: {data.get('message')}")
    return data


def _normalize_phone(p: str) -> str:
    """归一化手机号：剥除空格/+/-/括号，丢弃 86 国码前缀。"""
    if not p:
        return ""
    s = "".join(ch for ch in str(p) if ch.isdigit())
    if s.startswith("86") and len(s) == 13:
        s = s[2:]
    return s


@skill(registry=skill_registry, name="企名片.按手机号查投资人",
       version="1.0", timeout=10, retry=1, fallback=[])
async def qmingpian_search_person_by_phone(phone: str) -> list[dict]:
    """用手机号在企名片精准查投资人（open_id 鉴权）。
    手机号天然唯一，可绕开同名歧义。+86/空格/横杠自动归一化。
    没匹配返回 []。"""
    p = _normalize_phone(phone)
    if not p:
        return []
    async with httpx.AsyncClient() as client:
        resp = await client.post(
            f"{BASE_URL}/Person/searchPerson",
            data=_base({"keywords": p}),
        )
    body = _check(resp)
    items = body.get("data", {}).get("list", []) or []
    # 同 person_id 可能多次返回（多张名片），去重
    seen, out = set(), []
    for it in items:
        pid = it.get("person_id")
        if not pid or pid in seen:
            continue
        seen.add(pid)
        out.append(it)
    return out


@skill(registry=skill_registry, name="企名片.查询投资人",
       version="1.1", timeout=10, retry=2, fallback=[])
async def qmingpian_search_person(keywords: str) -> list[dict]:
    """企名片搜索投资人。

    用 team_uuid + unionid 鉴权（不是 open_id）—— 这样：
    1. 字段更全（含 zhiwu/icon/url/agency_id/case/style/is_develop）
    2. **每张名片对应一条记录**：同 person_id 有 N 张名片就返回 N 条，
       url 字段是该条对应的名片图。
       前端要"拿所有名片"就按 person_id 聚合 url。
    """
    data = {
        "keywords": keywords,
        "team_uuid": settings.qmingpian_team_uuid,
        "unionid": settings.qmingpian_unionid,
    }
    async with httpx.AsyncClient() as client:
        resp = await client.post(f"{BASE_URL}/Person/searchPerson", data=data)
    body = _check(resp)
    return body.get("data", {}).get("list", []) or []


@skill(registry=skill_registry, name="企名片.添加投资人",
       version="1.2", timeout=10, retry=1)
async def qmingpian_add_person(
    name: str,
    agency: str,
    phone: str = "",
    wechat: str = "",
    email: str = "",
    position: str = "",             # 企名片字段 zhiwu
    tags: list[str] | None = None,  # 企名片字段 tag（"|" 分隔）
    level: str = "",                # 投资人级别：'高' / '低'
    gender: str = "",               # 性别：'男' / '女'，企名片字段 sex
    office_location: str = "",      # 办公地区
    introduction: str = "",         # 简介
    is_dimission: int | None = None,  # 是否离职：1=离职 0=在职
    card_url: str = "",             # 名片图企名片 OSS URL（先调 qmingpian_upload_file 拿）
) -> dict:
    """新增投资人到企名片。

    注意：企名片侧的实际字段名是 zhiwu/sex/tag（不是 position/gender/tags），
    本函数在 Python 层保留更友好的命名并内部映射。

    返回 {person_id, ...}。传 card_url 可一并绑名片，无需再调 add_person_card。
    """
    form = _base({
        "name": name,
        "agency": agency,
    })
    if phone:
        form["phone"] = json.dumps([phone])
    if wechat:
        form["wechat"] = json.dumps([wechat])
    if email:
        form["email"] = json.dumps([email])
    if position:
        form["zhiwu"] = position          # 之前误用 'position'，企名片侧字段是 zhiwu
    if tags:
        form["tag"] = "|".join(tags)       # 之前误用 'content'，企名片侧字段是 tag
    if level:
        form["level"] = level
    if gender:
        form["sex"] = gender
    if office_location:
        form["office_location"] = office_location
    if introduction:
        form["introduction"] = introduction
    if is_dimission is not None:
        form["is_dimission"] = str(int(is_dimission))
    if card_url:
        form["card_url"] = card_url
    async with httpx.AsyncClient() as client:
        resp = await client.post(f"{BASE_URL}/Person/addPersonInfo", data=form)
    return _check(resp).get("data", {})


@skill(registry=skill_registry, name="企名片.更新投资人",
       version="1.0", timeout=10, retry=1)
async def qmingpian_edit_person(
    person_id: str,
    name: str = "",
    agency: str = "",
    phone: str = "",
    wechat: str = "",
    email: str = "",
    position: str = "",
) -> dict:
    form = _base({"person_id": person_id})
    if name:
        form["name"] = name
    if agency:
        form["agency"] = agency
    if phone:
        form["phone"] = json.dumps([phone])
    if wechat:
        form["wechat"] = json.dumps([wechat])
    if email:
        form["email"] = json.dumps([email])
    if position:
        form["zhiwu"] = position    # 企名片侧字段是 zhiwu，不是 position
    async with httpx.AsyncClient() as client:
        resp = await client.post(f"{BASE_URL}/Person/editPersonInfo", data=form)
    return _check(resp).get("data", {})


@skill(registry=skill_registry, name="企名片.绑定投资人名片",
       version="1.0", timeout=15, retry=1)
async def qmingpian_add_person_card(
    person_id: str,
    img_url: str,
    create_name: str,
) -> dict:
    """把 /Upload/file 拿到的 url 绑定到投资人记录，使企名片 PC 端能看到名片。

    注意：本接口不用 open_id，鉴权走 team_uuid+unionid。
    """
    if not settings.qmingpian_team_uuid or not settings.qmingpian_unionid:
        raise ValueError("缺少 QMINGPIAN_TEAM_UUID 或 QMINGPIAN_UNIONID")
    data = {
        "url": img_url,
        "person_id": person_id,
        "belong": "投资人库列表页",
        "team_uuid": settings.qmingpian_team_uuid,
        "unionid": settings.qmingpian_unionid,
        "create_name": create_name,
    }
    async with httpx.AsyncClient() as client:
        resp = await client.post(f"{BASE_URL}/Person/addPersonCard", data=data)
    return _check(resp).get("data", {})


@skill(registry=skill_registry, name="企名片.上传文件",
       version="1.0", timeout=60, retry=1)
async def qmingpian_upload_file(
    file_bytes: bytes,
    filename: str,
    mime_type: str = "application/octet-stream",
) -> dict:
    """上传任意文件到企名片侧 OSS（用于名片图等）。

    返回：{url, md5, size, file_name, ext}
    企名片实际返回结构是 data.list.{...}，内部已剥包装。
    """
    async with httpx.AsyncClient(timeout=60) as client:
        resp = await client.post(
            f"{BASE_URL}/Upload/file",
            data={"open_id": settings.qmingpian_token},
            files={"file": (filename, file_bytes, mime_type)},
        )
    data = _check(resp)
    return (data.get("data", {}) or {}).get("list", {}) or {}


@skill(registry=skill_registry, name="企名片.设置投资人熟悉度",
       version="1.0", timeout=10, retry=1)
async def qmingpian_add_familiar_person(
    name: str,
    agency: str,
    user_name: str,
    level: str,
) -> dict:
    """新建：某 IR 对某投资人的熟悉度（首次设置，无历史值）。
    参数名是 `name`（不是 person_name），这是企名片 addFamiliarPerson 的格式。
    """
    form = _base({
        "name": name,
        "agency": agency,
        "user_name": user_name,
        "level": level,
    })
    async with httpx.AsyncClient() as client:
        resp = await client.post(f"{BASE_URL}/Person/addFamiliarPerson", data=form)
    return _check(resp).get("data", {})


@skill(registry=skill_registry, name="企名片.编辑投资人熟悉度",
       version="1.0", timeout=10, retry=1)
async def qmingpian_update_familiar_person(
    name: str,
    agency: str,
    user_name: str,
    level: str,
) -> dict:
    """编辑：某 IR 对某投资人的熟悉度（已有历史值，覆盖）。
    注意：企名片 updateFamiliarPerson 的参数名是 `person_name`（不是 `name`），
    我们对外保持 `name` 参数命名一致，内部映射。
    """
    form = _base({
        "person_name": name,
        "agency": agency,
        "user_name": user_name,
        "level": level,
    })
    async with httpx.AsyncClient() as client:
        resp = await client.post(f"{BASE_URL}/Person/updateFamiliarPerson", data=form)
    return _check(resp).get("data", {})


@skill(registry=skill_registry, name="企名片.更新投资人标签",
       version="1.0", timeout=10, retry=1)
async def qmingpian_update_person_tags(
    name: str,
    agency: str,
    tags: list[str],
) -> dict:
    form = _base({
        "name": name,
        "agency": agency,
        "content": "|".join(tags),
    })
    async with httpx.AsyncClient() as client:
        resp = await client.post(f"{BASE_URL}/Person/updatePersonTag", data=form)
    return _check(resp).get("data", {})


@skill(registry=skill_registry, name="企名片.添加投资人纪要",
       version="1.0", timeout=10, retry=1)
async def qmingpian_add_person_summary(
    name: str,
    agency: str,
    summary: str,
    user_name: str,
) -> dict:
    form = _base({
        "name": name,
        "agency": agency,
        "summary": summary,
        "user_name": user_name,
    })
    async with httpx.AsyncClient() as client:
        resp = await client.post(f"{BASE_URL}/Summary/addPersonSummary", data=form)
    return _check(resp).get("data", {})


@skill(registry=skill_registry, name="企名片.添加机构纪要",
       version="1.1", timeout=10, retry=1)
async def qmingpian_add_agency_summary(
    agency: str,
    summary: str,
    user_name: str,
) -> dict:
    """企名片 addAgencySummary：注意机构名字段实际叫 `name`（不是 agency），
    跟 addPersonSummary 不一致。"""
    form = _base({
        "name": agency,
        "summary": summary,
        "user_name": user_name,
    })
    async with httpx.AsyncClient() as client:
        resp = await client.post(f"{BASE_URL}/Summary/addAgencySummary", data=form)
    return _check(resp).get("data", {})


@skill(registry=skill_registry, name="企名片.查询机构",
       version="1.0", timeout=10, retry=2, fallback=[])
async def qmingpian_search_agency(keywords: str, num: int = 20) -> list[dict]:
    """企名片多维机构库检索（searchAgency）。"""
    async with httpx.AsyncClient() as client:
        resp = await client.post(
            f"{BASE_URL}/Agency/searchAgency",
            data=_base({"keywords": keywords, "num": str(num)}),
        )
    data = _check(resp)
    return data.get("data", {}).get("list", [])


@skill(registry=skill_registry, name="企名片.检索外部机构",
       version="1.0", timeout=10, retry=2, fallback=[])
async def qmingpian_search_external_agency(keywords: str) -> list:
    """企名片外部机构库检索（agencyInfoList，又名 searchWaiBuAgency）。
    返回元素一般为字符串机构名（不是 dict），覆盖比 searchAgency 更广。
    """
    async with httpx.AsyncClient() as client:
        resp = await client.post(
            f"{BASE_URL}/Agency/agencyInfoList",
            data=_base({"keywords": keywords}),
        )
    data = _check(resp)
    return data.get("data", {}).get("list", []) or []


@skill(registry=skill_registry, name="企名片.添加机构",
       version="1.0", timeout=10, retry=1)
async def qmingpian_add_agency(name: str) -> dict:
    """企名片新增机构（addAgencyInfo）。
    幂等处理：「机构已存在」(status=1) 视为成功，返回 {existed: True}。
    """
    async with httpx.AsyncClient() as client:
        resp = await client.post(
            f"{BASE_URL}/Agency/addAgencyInfo",
            data=_base({"name": name}),
        )
    body = resp.json()
    if str(body.get("status")) == "0":
        return body.get("data", {}) or {}
    if str(body.get("message", "")).startswith("机构已存在"):
        return {"existed": True}
    raise ValueError(f"企名片 API error {body.get('status')}: {body.get('message')}")


@skill(registry=skill_registry, name="企名片.添加机构文件",
       version="1.0", timeout=15, retry=1)
async def qmingpian_add_agency_file(
    agency_name: str,
    filename: str,
    file_url: str,
    user_name: str = "",
) -> dict:
    """企名片为某机构挂载文件（addAgencyFile，如 BP/DataPack/Term Sheet 等）。

    参数：
    - agency_name: 机构名
    - filename: 文件显示名（带扩展名）
    - file_url: 文件公网可访问 URL（如先调 /Upload/file 拿到的 URL）
    - user_name: IR 在企名片侧用户名（create_name 同语义）

    内部会先调 addAgencyInfo 确保机构存在，再调 addAgencyFile。
    """
    # 先确保机构存在；失败也继续（多数是已存在）
    try:
        await qmingpian_add_agency(agency_name)
    except Exception:
        pass
    async with httpx.AsyncClient(timeout=15) as client:
        resp = await client.post(
            f"{BASE_URL}/ProductFile/addAgencyFile",
            data=_base({
                "name": agency_name,
                "file_name": filename,
                "url": file_url,
                "user_name": user_name,
            }),
        )
    return _check(resp).get("data", {})


@skill(registry=skill_registry, name="企名片.搜索企名片机构",
       version="1.0", timeout=10, retry=2, fallback=[])
async def qmingpian_search_jigou(keywords: str) -> list[dict]:
    async with httpx.AsyncClient() as client:
        resp = await client.post(
            f"{BASE_URL}/Search/searchJigou",
            data=_base({"keywords": keywords}),
        )
    data = _check(resp)
    return data.get("data", {}).get("list", [])


def _agency_brand_for_match(s: str) -> str:
    """同 direct.py 的逻辑：去掉常见地名前缀，取前 2 字。仅供本模块 export 解析时匹配。"""
    s = (s or "").strip()
    if not s:
        return ""
    prefixes = ["珠海", "上海", "北京", "深圳", "广州", "杭州", "成都", "南京", "苏州",
                "天津", "重庆", "宁波", "厦门", "西安", "武汉", "长沙", "香港", "澳门",
                "广东省", "浙江省", "江苏省", "山东省", "福建省"]
    for pre in sorted(prefixes, key=len, reverse=True):
        if s.startswith(pre):
            s = s[len(pre):]
            break
    return s[:2]


def _same_agency_loose(a: str, b: str) -> bool:
    a, b = (a or "").strip(), (b or "").strip()
    if not a or not b:
        return False
    if a == b:
        return True
    if len(a) >= 2 and a in b:
        return True
    if len(b) >= 2 and b in a:
        return True
    ba, bb = _agency_brand_for_match(a), _agency_brand_for_match(b)
    return bool(ba) and ba == bb


@skill(registry=skill_registry, name="企名片.导出投资人详情",
       version="1.2", timeout=15, retry=1)
async def qmingpian_export_person(person_name: str = "", expected_agency: str = "",
                                  person_id: str = "") -> dict:
    """导出投资人详情（xlsx）。

    优先使用 **person_id** 精准定位（推荐）—— 避免同名歧义，返回的 xlsx 总是单条记录的
    完整 sheet 集（详情 + 纪要 + 历史 + 熟悉人）。
    fallback：仅 person_name → 同名场景下"投资人详情"会有多行，按 expected_agency
    fuzzy 选；其他 sheet 可能被企名片 API 省略。

    返回：
    - agency / phone / email / industry  —— 来自"投资人详情"
    - summaries / history / familiar_persons —— 对应 sheet
    """
    import io
    from openpyxl import load_workbook

    if not person_id and not person_name:
        raise ValueError("person_id 和 person_name 至少给一个")

    payload: dict = {}
    if person_id:
        payload["person_id"] = person_id
    if person_name:
        payload["person_name"] = person_name

    async with httpx.AsyncClient(timeout=15) as client:
        resp = await client.post(
            f"{BASE_URL}/Export/exportPersonOpen",
            data=_base(payload),
        )

    ct = resp.headers.get("content-type", "")
    if not ct.startswith("application/vnd.openxmlformats"):
        try:
            d = resp.json()
            raise ValueError(f"企名片 export error: status={d.get('status')} message={d.get('message')}")
        except Exception:
            raise ValueError(f"企名片 export 失败，返回非 xlsx: {ct}")

    wb = load_workbook(io.BytesIO(resp.content), read_only=True)
    result: dict = {"summaries": [], "history": [], "familiar_persons": []}

    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 3:
            continue
        headers = [str(c).strip() if c else "" for c in rows[1]]

        if sn == "投资人详情":
            # 多行：每行一个同名人。按 expected_agency fuzzy 选；不匹配 → 取第一行兜底
            data_rows = rows[2:]
            chosen_row = None
            if expected_agency and "机构" in headers:
                idx_ag = headers.index("机构")
                for row in data_rows:
                    if not row or len(row) <= idx_ag:
                        continue
                    row_agency = str(row[idx_ag] or "").strip()
                    if _same_agency_loose(row_agency, expected_agency):
                        chosen_row = row
                        break
            if chosen_row is None and data_rows:
                chosen_row = data_rows[0]
            if chosen_row is None:
                continue
            for h, v in zip(headers, chosen_row):
                if not h or v in (None, ""):
                    continue
                if h == "机构":
                    result["agency"] = v
                elif h == "手机":
                    result["phone"] = [str(v)]
                elif h == "邮箱":
                    result["email"] = [str(v)]
                elif h in ("FAwork行业", "行业"):
                    result["industry"] = str(v)

        elif sn == "投资人纪要":
            # 表头: 纪要内容 / 创建人 / 创建时间
            for row in rows[2:]:
                if not row or all(c in (None, "") for c in row):
                    continue
                result["summaries"].append({
                    "content": str(row[0]) if len(row) > 0 and row[0] else "",
                    "creator": str(row[1]) if len(row) > 1 and row[1] else "",
                    "created_at": str(row[2]) if len(row) > 2 and row[2] else "",
                })

        elif sn == "历史推荐":
            # 表头: 事件名 / 机构 / 行业 / 服务轮次 / 状态 / 反馈及进展 / 对接时间
            for row in rows[2:]:
                if not row or all(c in (None, "") for c in row):
                    continue
                result["history"].append({
                    "event": str(row[0]) if len(row) > 0 and row[0] else "",
                    "agency": str(row[1]) if len(row) > 1 and row[1] else "",
                    "industry": str(row[2]) if len(row) > 2 and row[2] else "",
                    "round": str(row[3]) if len(row) > 3 and row[3] else "",
                    "status": str(row[4]) if len(row) > 4 and row[4] else "",
                    "feedback": str(row[5]) if len(row) > 5 and row[5] else "",
                    "contact_time": str(row[6]) if len(row) > 6 and row[6] else "",
                })

        elif sn == "熟悉人":
            # 表头: 姓名 / 熟悉度
            for row in rows[2:]:
                if not row or all(c in (None, "") for c in row):
                    continue
                name = str(row[0]) if len(row) > 0 and row[0] else ""
                level = str(row[1]) if len(row) > 1 and row[1] else ""
                if name:
                    result["familiar_persons"].append({"name": name, "level": level})

    return result


@skill(registry=skill_registry, name="企名片.导出机构详情",
       version="1.0", timeout=15, retry=1)
async def qmingpian_export_agency(agency_name: str) -> dict:
    """导出机构详情（xlsx）。

    返回 3 个 sheet：
    - 机构详情 → tags
    - 机构纪要 → summaries: list of {content, creator, created_at}
    - 历史推荐 → history: list of {event, industry, round, status, feedback,
                                   recommended_investor, contact_time}
    """
    import io
    from openpyxl import load_workbook

    if not agency_name:
        raise ValueError("agency_name 不能为空")

    async with httpx.AsyncClient(timeout=15) as client:
        resp = await client.post(
            f"{BASE_URL}/Export/exportAgencyOpen",
            data=_base({"agency_name": agency_name}),
        )

    ct = resp.headers.get("content-type", "")
    if not ct.startswith("application/vnd.openxmlformats"):
        try:
            d = resp.json()
            raise ValueError(f"企名片 agency export error: status={d.get('status')} message={d.get('message')}")
        except Exception:
            raise ValueError(f"企名片 agency export 失败，返回非 xlsx: {ct}")

    wb = load_workbook(io.BytesIO(resp.content), read_only=True)
    result: dict = {"tags": [], "summaries": [], "history": []}

    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        if sn == "机构详情":
            # row 1 = header（机构标签）, row 2 = value（"|" 或空格分隔，或多列）
            if len(rows) >= 3 and rows[2]:
                vals = []
                for c in rows[2]:
                    if c:
                        vals.append(str(c).strip())
                if vals:
                    # 可能是单个 cell 用分隔符
                    if len(vals) == 1 and ("|" in vals[0] or "," in vals[0]):
                        result["tags"] = [t.strip() for t in vals[0].replace(",", "|").split("|") if t.strip()]
                    else:
                        result["tags"] = vals

        elif sn == "机构纪要":
            for row in rows[2:]:
                if not row or all(c in (None, "") for c in row):
                    continue
                result["summaries"].append({
                    "content": str(row[0]) if len(row) > 0 and row[0] else "",
                    "creator": str(row[1]) if len(row) > 1 and row[1] else "",
                    "created_at": str(row[2]) if len(row) > 2 and row[2] else "",
                })

        elif sn == "历史推荐":
            for row in rows[2:]:
                if not row or all(c in (None, "") for c in row):
                    continue
                result["history"].append({
                    "event": str(row[0]) if len(row) > 0 and row[0] else "",
                    "industry": str(row[1]) if len(row) > 1 and row[1] else "",
                    "round": str(row[2]) if len(row) > 2 and row[2] else "",
                    "status": str(row[3]) if len(row) > 3 and row[3] else "",
                    "feedback": str(row[4]) if len(row) > 4 and row[4] else "",
                    "recommended_investor": str(row[5]) if len(row) > 5 and row[5] else "",
                    "contact_time": str(row[6]) if len(row) > 6 and row[6] else "",
                })

    return result


@skill(registry=skill_registry, name="企名片.导出ongoing项目对接",
       version="1.0", timeout=15, retry=1)
async def qmingpian_export_ongoing_lunci(event_name: str = "") -> dict:
    """导出 ongoing 项目对接进展（机构 + 对接投资人）。

    - event_name="" → 返回所有 ongoing 项目涉及的机构/投资人（全量）
    - event_name="珀乐互动/A轮/3000万" → 返回该项目的所有对接清单

    返回：{contacts: list[{agency, person}], count: int}
    """
    import io
    from openpyxl import load_workbook
    async with httpx.AsyncClient(timeout=15) as client:
        resp = await client.post(
            f"{BASE_URL}/Export/exportLunciOpen",
            data=_base({"event_name": event_name}),
        )
    ct = resp.headers.get("content-type", "")
    if not ct.startswith("application/vnd.openxmlformats"):
        try:
            d = resp.json()
            raise ValueError(f"企名片 lunci export error: status={d.get('status')} message={d.get('message')}")
        except Exception:
            raise ValueError(f"企名片 lunci export 失败，返回非 xlsx: {ct}")
    wb = load_workbook(io.BytesIO(resp.content), read_only=True)
    contacts: list[dict] = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 3:
            continue
        # row 0: title, row 1: header (机构名 / 投资人名), row 2+: data
        for row in rows[2:]:
            if not row or all(c in (None, "") for c in row):
                continue
            agency = str(row[0]) if len(row) > 0 and row[0] else ""
            person = str(row[1]) if len(row) > 1 and row[1] else ""
            if agency or person:
                contacts.append({"agency": agency, "person": person})
    return {"contacts": contacts, "count": len(contacts)}


# 旧 stub 已废弃 —— 完整实现见上方 qmingpian_export_agency(agency_name)
